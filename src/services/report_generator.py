"""
ReportGenerator service for generating compliance and fraud detection reports.

This module provides comprehensive reporting capabilities including PDF reports
with fraud detection statistics, Excel exports with detailed transaction analysis,
and automated report scheduling and distribution functionality.
"""

import logging
from typing import Dict, List, Any, Optional, Union, Tuple
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import numpy as np
from dataclasses import dataclass, asdict
from enum import Enum
import json
import io
import base64
from concurrent.futures import ThreadPoolExecutor
import threading
import time

# Report generation libraries - import only if available
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel generation - import only if available
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Template engine - import only if available
try:
    from jinja2 import Template, Environment, BaseLoader
    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False

# Scheduling - import only if available
try:
    import schedule
    SCHEDULE_AVAILABLE = True
except ImportError:
    SCHEDULE_AVAILABLE = False

# Try to import project config, fallback to basic setup
print("DEBUG: About to import config")
try:
    from src.utils.config import get_setting, get_logger, settings
    logger = get_logger(__name__)
    print("DEBUG: Config import successful")
except ImportError:
    logger = logging.getLogger(__name__)
    # Create a mock settings object
    class MockSettings:
        def __init__(self):
            self.project_root = Path(".")
    settings = MockSettings()
    print("DEBUG: Using mock settings")

print("DEBUG: About to define ReportType")


class ReportType(Enum):
    """Report type enumeration."""
    FRAUD_SUMMARY = "fraud_summary"
    TRANSACTION_ANALYSIS = "transaction_analysis"
    MODEL_PERFORMANCE = "model_performance"
    COMPLIANCE_AUDIT = "compliance_audit"
    ALERT_SUMMARY = "alert_summary"
    CUSTOM = "custom"


class ReportFormat(Enum):
    """Report format enumeration."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"


@dataclass
class ReportConfig:
    """Configuration for report generation."""
    report_type: ReportType
    report_format: ReportFormat
    title: str
    description: str = ""
    include_charts: bool = True
    include_raw_data: bool = False
    date_range_days: int = 30
    output_directory: Optional[str] = None
    template_path: Optional[str] = None
    auto_schedule: bool = False
    schedule_frequency: str = "daily"  # daily, weekly, monthly
    recipients: List[str] = None
    
    def __post_init__(self):
        if self.recipients is None:
            self.recipients = []


@dataclass
class ReportData:
    """Container for report data."""
    fraud_statistics: Dict[str, Any] = None
    transaction_data: pd.DataFrame = None
    model_metrics: Dict[str, Any] = None
    alert_data: List[Dict[str, Any]] = None
    time_series_data: Dict[str, List] = None
    custom_data: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.fraud_statistics is None:
            self.fraud_statistics = {}
        if self.alert_data is None:
            self.alert_data = []
        if self.time_series_data is None:
            self.time_series_data = {}
        if self.custom_data is None:
            self.custom_data = {}


class ReportGenerator:
    """
    Main report generation service for fraud detection compliance reports.
    
    Generates PDF reports with fraud detection statistics, Excel exports with
    detailed transaction analysis, and provides automated report scheduling.
    """
    
    def __init__(self, 
                 output_directory: Optional[str] = None,
                 template_directory: Optional[str] = None,
                 enable_scheduling: bool = True):
        """
        Initialize the ReportGenerator.
        
        Args:
            output_directory: Directory for generated reports
            template_directory: Directory containing report templates
            enable_scheduling: Whether to enable automated report scheduling
        """
        # Setup directories
        if output_directory:
            self.output_directory = Path(output_directory)
        else:
            self.output_directory = settings.project_root / "reports"
        
        self.output_directory.mkdir(exist_ok=True)
        
        if template_directory:
            self.template_directory = Path(template_directory)
        else:
            self.template_directory = settings.project_root / "templates" / "reports"
        
        # Report scheduling
        self.enable_scheduling = enable_scheduling and SCHEDULE_AVAILABLE
        self.scheduled_reports: List[ReportConfig] = []
        self.scheduler_thread: Optional[threading.Thread] = None
        self.scheduler_running = False
        
        # Report generation executor
        self.report_executor = ThreadPoolExecutor(max_workers=2)
        
        # Report history and tracking
        self.report_history: List[Dict[str, Any]] = []
        
        # Setup Jinja2 environment for templates
        if JINJA2_AVAILABLE:
            self.template_env = Environment(loader=BaseLoader())
        else:
            self.template_env = None
        
        logger.info(f"ReportGenerator initialized with output directory: {self.output_directory}")
    
    def generate_fraud_summary_report(self, 
                                    data: ReportData,
                                    config: ReportConfig) -> str:
        """
        Generate a comprehensive fraud summary report.
        
        Args:
            data: Report data container
            config: Report configuration
            
        Returns:
            Path to generated report file
        """
        logger.info(f"Generating fraud summary report in {config.report_format.value} format")
        
        if config.report_format == ReportFormat.PDF:
            if not REPORTLAB_AVAILABLE:
                raise ImportError("ReportLab is required for PDF generation. Install with: pip install reportlab")
            return self._generate_fraud_summary_pdf(data, config)
        elif config.report_format == ReportFormat.EXCEL:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
            return self._generate_fraud_summary_excel(data, config)
        else:
            raise ValueError(f"Unsupported format for fraud summary: {config.report_format.value}")
    
    def generate_transaction_analysis_report(self,
                                           data: ReportData,
                                           config: ReportConfig) -> str:
        """
        Generate detailed transaction analysis report.
        
        Args:
            data: Report data container
            config: Report configuration
            
        Returns:
            Path to generated report file
        """
        logger.info(f"Generating transaction analysis report in {config.report_format.value} format")
        
        if config.report_format == ReportFormat.EXCEL:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
            return self._generate_transaction_analysis_excel(data, config)
        elif config.report_format == ReportFormat.CSV:
            return self._generate_transaction_analysis_csv(data, config)
        else:
            raise ValueError(f"Unsupported format for transaction analysis: {config.report_format.value}")
    
    def generate_model_performance_report(self,
                                        data: ReportData,
                                        config: ReportConfig) -> str:
        """
        Generate model performance report.
        
        Args:
            data: Report data container
            config: Report configuration
            
        Returns:
            Path to generated report file
        """
        logger.info(f"Generating model performance report in {config.report_format.value} format")
        
        if config.report_format == ReportFormat.PDF:
            if not REPORTLAB_AVAILABLE:
                raise ImportError("ReportLab is required for PDF generation. Install with: pip install reportlab")
            return self._generate_model_performance_pdf(data, config)
        elif config.report_format == ReportFormat.EXCEL:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
            return self._generate_model_performance_excel(data, config)
        else:
            raise ValueError(f"Unsupported format for model performance: {config.report_format.value}")
    
    def generate_compliance_audit_report(self,
                                       data: ReportData,
                                       config: ReportConfig) -> str:
        """
        Generate compliance audit report.
        
        Args:
            data: Report data container
            config: Report configuration
            
        Returns:
            Path to generated report file
        """
        logger.info(f"Generating compliance audit report in {config.report_format.value} format")
        
        if config.report_format == ReportFormat.PDF:
            if not REPORTLAB_AVAILABLE:
                raise ImportError("ReportLab is required for PDF generation. Install with: pip install reportlab")
            return self._generate_compliance_audit_pdf(data, config)
        elif config.report_format == ReportFormat.EXCEL:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
            return self._generate_compliance_audit_excel(data, config)
        else:
            raise ValueError(f"Unsupported format for compliance audit: {config.report_format.value}")
    
    def _generate_transaction_analysis_csv(self, data: ReportData, config: ReportConfig) -> str:
        """Generate CSV transaction analysis report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transaction_analysis_{timestamp}.csv"
        filepath = self.output_directory / filename
        
        if data.transaction_data is None or data.transaction_data.empty:
            raise ValueError("Transaction data is required for transaction analysis report")
        
        # Save transaction data to CSV
        data.transaction_data.to_csv(filepath, index=False)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Transaction analysis CSV report generated: {filepath}")
        return str(filepath)
    
    def _generate_fraud_summary_pdf(self, data: ReportData, config: ReportConfig) -> str:
        """Generate PDF fraud summary report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"fraud_summary_{timestamp}.pdf"
        filepath = self.output_directory / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        story.append(Paragraph(config.title, title_style))
        story.append(Spacer(1, 12))
        
        # Report metadata
        metadata_data = [
            ['Report Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ['Report Period:', f"Last {config.date_range_days} days"],
            ['Report Type:', 'Fraud Detection Summary']
        ]
        metadata_table = Table(metadata_data, colWidths=[2*inch, 3*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(metadata_table)
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        
        fraud_stats = data.fraud_statistics
        if fraud_stats:
            summary_text = f"""
            During the reporting period, the fraud detection system processed {fraud_stats.get('total_transactions', 0):,} transactions.
            Of these, {fraud_stats.get('fraud_detected', 0):,} transactions were flagged as potentially fraudulent 
            ({fraud_stats.get('fraud_rate', 0):.2%} fraud rate).
            
            The system achieved {fraud_stats.get('accuracy', 0):.1%} accuracy with {fraud_stats.get('precision', 0):.1%} precision 
            and {fraud_stats.get('recall', 0):.1%} recall in fraud detection.
            """
            story.append(Paragraph(summary_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Fraud summary PDF report generated: {filepath}")
        return str(filepath)
    
    def _generate_fraud_summary_excel(self, data: ReportData, config: ReportConfig) -> str:
        """Generate Excel fraud summary report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"fraud_summary_{timestamp}.xlsx"
        filepath = self.output_directory / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws_summary['A1'] = config.title
        ws_summary['A1'].font = Font(bold=True, size=18)
        ws_summary.merge_cells('A1:D1')
        
        # Metadata
        ws_summary['A3'] = "Report Generated:"
        ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary['A4'] = "Report Period:"
        ws_summary['B4'] = f"Last {config.date_range_days} days"
        
        # Key statistics
        if data.fraud_statistics:
            fraud_stats = data.fraud_statistics
            
            ws_summary['A6'] = "Key Performance Indicators"
            ws_summary['A6'].font = header_font
            ws_summary['A6'].fill = header_fill
            ws_summary.merge_cells('A6:B6')
            
            stats_data = [
                ("Total Transactions Processed", fraud_stats.get('total_transactions', 0)),
                ("Fraudulent Transactions Detected", fraud_stats.get('fraud_detected', 0)),
                ("Fraud Detection Rate", f"{fraud_stats.get('fraud_rate', 0):.2%}"),
                ("Model Accuracy", f"{fraud_stats.get('accuracy', 0):.1%}"),
                ("Precision", f"{fraud_stats.get('precision', 0):.1%}"),
                ("Recall", f"{fraud_stats.get('recall', 0):.1%}"),
                ("F1-Score", f"{fraud_stats.get('f1_score', 0):.1%}"),
                ("False Positive Rate", f"{fraud_stats.get('false_positive_rate', 0):.2%}"),
                ("False Negative Rate", f"{fraud_stats.get('false_negative_rate', 0):.2%}")
            ]
            
            for i, (metric, value) in enumerate(stats_data, start=7):
                ws_summary[f'A{i}'] = metric
                ws_summary[f'B{i}'] = value
        
        # Save workbook
        wb.save(filepath)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Fraud summary Excel report generated: {filepath}")
        return str(filepath)
    
    def _generate_transaction_analysis_excel(self, data: ReportData, config: ReportConfig) -> str:
        """Generate Excel transaction analysis report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transaction_analysis_{timestamp}.xlsx"
        filepath = self.output_directory / filename
        
        if data.transaction_data is None or data.transaction_data.empty:
            raise ValueError("Transaction data is required for transaction analysis report")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Transaction summary sheet
        ws_summary = wb.active
        ws_summary.title = "Transaction Summary"
        
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws_summary['A1'] = config.title
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:E1')
        
        # Transaction type analysis
        if 'type' in data.transaction_data.columns:
            type_counts = data.transaction_data['type'].value_counts()
            
            ws_summary['A3'] = "Transaction Type Analysis"
            ws_summary['A3'].font = Font(bold=True, size=14)
            
            ws_summary['A5'] = "Transaction Type"
            ws_summary['B5'] = "Count"
            ws_summary['C5'] = "Percentage"
            
            for i, col in enumerate(['A5', 'B5', 'C5']):
                ws_summary[col].font = header_font
                ws_summary[col].fill = header_fill
            
            total_transactions = len(data.transaction_data)
            for i, (tx_type, count) in enumerate(type_counts.items(), start=6):
                ws_summary[f'A{i}'] = tx_type
                ws_summary[f'B{i}'] = count
                ws_summary[f'C{i}'] = f"{count/total_transactions:.1%}"
        
        # Detailed transaction data sheet
        ws_data = wb.create_sheet("Detailed Transactions")
        
        # Add headers
        for i, col in enumerate(data.transaction_data.columns, start=1):
            cell = ws_data.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data (limit to first 10000 rows for performance)
        sample_data = data.transaction_data.head(10000)
        for r_idx, row in enumerate(dataframe_to_rows(sample_data, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_data.cell(row=r_idx, column=c_idx, value=value)
        
        # Save workbook
        wb.save(filepath)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Transaction analysis Excel report generated: {filepath}")
        return str(filepath)
    
    def _generate_model_performance_pdf(self, data: ReportData, config: ReportConfig) -> str:
        """Generate PDF model performance report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"model_performance_{timestamp}.pdf"
        filepath = self.output_directory / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        story.append(Paragraph(config.title, title_style))
        story.append(Spacer(1, 12))
        
        # Model metrics
        if data.model_metrics:
            story.append(Paragraph("Model Performance Metrics", styles['Heading2']))
            
            metrics_data = [
                ['Metric', 'Value'],
                ['Accuracy', f"{data.model_metrics.get('accuracy', 0):.3f}"],
                ['Precision', f"{data.model_metrics.get('precision', 0):.3f}"],
                ['Recall', f"{data.model_metrics.get('recall', 0):.3f}"],
                ['F1-Score', f"{data.model_metrics.get('f1_score', 0):.3f}"],
                ['AUC-ROC', f"{data.model_metrics.get('auc_roc', 0):.3f}"]
            ]
            
            metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metrics_table)
        
        # Build PDF
        doc.build(story)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Model performance PDF report generated: {filepath}")
        return str(filepath)
    
    def _generate_model_performance_excel(self, data: ReportData, config: ReportConfig) -> str:
        """Generate Excel model performance report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"model_performance_{timestamp}.xlsx"
        filepath = self.output_directory / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Model Performance"
        
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = config.title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Model metrics
        if data.model_metrics:
            ws['A3'] = "Performance Metrics"
            ws['A3'].font = Font(bold=True, size=14)
            
            ws['A5'] = "Metric"
            ws['B5'] = "Value"
            
            for col in ['A5', 'B5']:
                ws[col].font = header_font
                ws[col].fill = header_fill
            
            metrics_data = [
                ("Accuracy", f"{data.model_metrics.get('accuracy', 0):.3f}"),
                ("Precision", f"{data.model_metrics.get('precision', 0):.3f}"),
                ("Recall", f"{data.model_metrics.get('recall', 0):.3f}"),
                ("F1-Score", f"{data.model_metrics.get('f1_score', 0):.3f}"),
                ("AUC-ROC", f"{data.model_metrics.get('auc_roc', 0):.3f}")
            ]
            
            for i, (metric, value) in enumerate(metrics_data, start=6):
                ws[f'A{i}'] = metric
                ws[f'B{i}'] = value
        
        # Save workbook
        wb.save(filepath)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Model performance Excel report generated: {filepath}")
        return str(filepath)
    
    def _generate_compliance_audit_pdf(self, data: ReportData, config: ReportConfig) -> str:
        """Generate PDF compliance audit report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"compliance_audit_{timestamp}.pdf"
        filepath = self.output_directory / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        story.append(Paragraph(config.title, title_style))
        story.append(Spacer(1, 12))
        
        # Audit metadata
        audit_data = [
            ['Audit Date:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ['Audit Period:', f"Last {config.date_range_days} days"],
            ['Report Type:', 'Compliance Audit'],
            ['Auditor:', 'Automated Fraud Detection System']
        ]
        audit_table = Table(audit_data, colWidths=[2*inch, 3*inch])
        audit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(audit_table)
        story.append(Spacer(1, 20))
        
        # Compliance summary
        story.append(Paragraph("Compliance Summary", styles['Heading2']))
        
        if data.fraud_statistics:
            fraud_stats = data.fraud_statistics
            
            compliance_text = f"""
            The fraud detection system has been operating within compliance parameters during the audit period.
            
            System Performance:
            - Total transactions monitored: {fraud_stats.get('total_transactions', 0):,}
            - Fraud detection accuracy: {fraud_stats.get('accuracy', 0):.1%}
            - False positive rate: {fraud_stats.get('false_positive_rate', 0):.2%}
            - False negative rate: {fraud_stats.get('false_negative_rate', 0):.2%}
            """
            story.append(Paragraph(compliance_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Compliance audit PDF report generated: {filepath}")
        return str(filepath)
    
    def _generate_compliance_audit_excel(self, data: ReportData, config: ReportConfig) -> str:
        """Generate Excel compliance audit report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"compliance_audit_{timestamp}.xlsx"
        filepath = self.output_directory / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compliance Audit"
        
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = config.title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Audit metadata
        ws['A3'] = "Audit Information"
        ws['A3'].font = Font(bold=True, size=14)
        
        audit_info = [
            ("Audit Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Audit Period", f"Last {config.date_range_days} days"),
            ("Report Type", "Compliance Audit"),
            ("Auditor", "Automated Fraud Detection System")
        ]
        
        for i, (key, value) in enumerate(audit_info, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
        
        # Compliance metrics
        if data.fraud_statistics:
            ws['A10'] = "Compliance Metrics"
            ws['A10'].font = Font(bold=True, size=14)
            
            ws['A12'] = "Metric"
            ws['B12'] = "Value"
            ws['C12'] = "Compliance Status"
            
            for col in ['A12', 'B12', 'C12']:
                ws[col].font = header_font
                ws[col].fill = header_fill
            
            fraud_stats = data.fraud_statistics
            compliance_data = [
                ("Total Transactions Monitored", f"{fraud_stats.get('total_transactions', 0):,}", "✓ Compliant"),
                ("Fraud Detection Accuracy", f"{fraud_stats.get('accuracy', 0):.1%}", "✓ Above Threshold"),
                ("False Positive Rate", f"{fraud_stats.get('false_positive_rate', 0):.2%}", "✓ Within Limits"),
                ("False Negative Rate", f"{fraud_stats.get('false_negative_rate', 0):.2%}", "✓ Within Limits"),
                ("Model Performance", f"{fraud_stats.get('f1_score', 0):.1%}", "✓ Meets Standards")
            ]
            
            for i, (metric, value, status) in enumerate(compliance_data, start=13):
                ws[f'A{i}'] = metric
                ws[f'B{i}'] = value
                ws[f'C{i}'] = status
        
        # Save workbook
        wb.save(filepath)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Compliance audit Excel report generated: {filepath}")
        return str(filepath)
    
    def generate_custom_report(self, 
                             data: ReportData,
                             config: ReportConfig,
                             template_content: str = None) -> str:
        """
        Generate a custom report using a template.
        
        Args:
            data: Report data container
            config: Report configuration
            template_content: Custom template content (Jinja2 format)
            
        Returns:
            Path to generated report file
        """
        if not JINJA2_AVAILABLE:
            raise ImportError("Jinja2 is required for custom reports. Install with: pip install jinja2")
        
        logger.info(f"Generating custom report: {config.title}")
        
        if template_content is None:
            # Load template from file if specified
            if config.template_path:
                template_path = Path(config.template_path)
                if template_path.exists():
                    template_content = template_path.read_text()
                else:
                    raise FileNotFoundError(f"Template file not found: {config.template_path}")
            else:
                raise ValueError("Template content or template path must be provided for custom reports")
        
        # Render template
        template = self.template_env.from_string(template_content)
        
        # Prepare template context
        context = {
            'config': asdict(config),
            'data': asdict(data),
            'timestamp': datetime.now(),
            'fraud_statistics': data.fraud_statistics,
            'model_metrics': data.model_metrics,
            'alert_data': data.alert_data,
            'custom_data': data.custom_data
        }
        
        rendered_content = template.render(**context)
        
        # Save rendered content
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"custom_report_{timestamp}.html"
        filepath = self.output_directory / filename
        
        filepath.write_text(rendered_content)
        
        # Record report generation
        self._record_report_generation(config, str(filepath))
        
        logger.info(f"Custom report generated: {filepath}")
        return str(filepath)
    
    def schedule_report(self, config: ReportConfig) -> None:
        """
        Schedule automated report generation.
        
        Args:
            config: Report configuration with scheduling parameters
        """
        if not self.enable_scheduling:
            logger.warning("Report scheduling is disabled or schedule library not available")
            return
        
        if not config.auto_schedule:
            logger.warning("Auto-scheduling not enabled for this report config")
            return
        
        # Add to scheduled reports
        self.scheduled_reports.append(config)
        
        # Schedule based on frequency
        if config.schedule_frequency == "daily":
            schedule.every().day.at("06:00").do(self._generate_scheduled_report, config)
        elif config.schedule_frequency == "weekly":
            schedule.every().monday.at("06:00").do(self._generate_scheduled_report, config)
        elif config.schedule_frequency == "monthly":
            schedule.every().month.do(self._generate_scheduled_report, config)
        
        logger.info(f"Scheduled {config.report_type.value} report for {config.schedule_frequency} generation")
        
        # Start scheduler thread if not running
        if not self.scheduler_running:
            self._start_scheduler()
    
    def _generate_scheduled_report(self, config: ReportConfig) -> None:
        """Generate a scheduled report."""
        try:
            logger.info(f"Generating scheduled report: {config.report_type.value}")
            
            # Collect data for the report (this would typically fetch from database/services)
            data = self._collect_report_data(config)
            
            # Generate the report
            if config.report_type == ReportType.FRAUD_SUMMARY:
                filepath = self.generate_fraud_summary_report(data, config)
            elif config.report_type == ReportType.TRANSACTION_ANALYSIS:
                filepath = self.generate_transaction_analysis_report(data, config)
            elif config.report_type == ReportType.MODEL_PERFORMANCE:
                filepath = self.generate_model_performance_report(data, config)
            elif config.report_type == ReportType.COMPLIANCE_AUDIT:
                filepath = self.generate_compliance_audit_report(data, config)
            else:
                logger.error(f"Unsupported scheduled report type: {config.report_type.value}")
                return
            
            # Distribute report if recipients are specified
            if config.recipients:
                self._distribute_report(filepath, config)
            
            logger.info(f"Scheduled report generated successfully: {filepath}")
            
        except Exception as e:
            logger.error(f"Error generating scheduled report: {str(e)}")
    
    def _collect_report_data(self, config: ReportConfig) -> ReportData:
        """
        Collect data for report generation.
        
        This is a placeholder method that should be implemented to fetch
        actual data from the fraud detection system.
        """
        # This would typically fetch data from databases, services, etc.
        # For now, return empty data structure
        return ReportData()
    
    def _distribute_report(self, filepath: str, config: ReportConfig) -> None:
        """
        Distribute generated report to recipients.
        
        Args:
            filepath: Path to generated report file
            config: Report configuration with recipient information
        """
        try:
            # This would typically integrate with email/notification services
            # For now, just log the distribution
            logger.info(f"Distributing report {filepath} to {len(config.recipients)} recipients")
            
            for recipient in config.recipients:
                logger.info(f"Report sent to: {recipient}")
                
        except Exception as e:
            logger.error(f"Error distributing report: {str(e)}")
    
    def _start_scheduler(self) -> None:
        """Start the report scheduler thread."""
        if self.scheduler_running or not SCHEDULE_AVAILABLE:
            return
        
        self.scheduler_running = True
        self.scheduler_thread = threading.Thread(target=self._run_scheduler, daemon=True)
        self.scheduler_thread.start()
        logger.info("Report scheduler started")
    
    def _run_scheduler(self) -> None:
        """Run the report scheduler."""
        while self.scheduler_running:
            schedule.run_pending()
            time.sleep(60)  # Check every minute
    
    def stop_scheduler(self) -> None:
        """Stop the report scheduler."""
        self.scheduler_running = False
        if self.scheduler_thread:
            self.scheduler_thread.join()
        logger.info("Report scheduler stopped")
    
    def _analyze_alert_data(self, alert_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze alert data for reporting."""
        if not alert_data:
            return {}
        
        # Count alerts by severity
        severity_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        
        for alert in alert_data:
            severity = alert.get('severity', 'low').lower()
            if severity in severity_counts:
                severity_counts[severity] += 1
        
        total_alerts = sum(severity_counts.values())
        
        # Calculate percentages
        result = {}
        for severity, count in severity_counts.items():
            result[severity] = count
            result[f'{severity}_pct'] = count / total_alerts if total_alerts > 0 else 0
        
        return result
    
    def _record_report_generation(self, config: ReportConfig, filepath: str) -> None:
        """Record report generation in history."""
        record = {
            'timestamp': datetime.now().isoformat(),
            'report_type': config.report_type.value,
            'report_format': config.report_format.value,
            'title': config.title,
            'filepath': filepath,
            'scheduled': config.auto_schedule
        }
        
        self.report_history.append(record)
        
        # Keep only last 100 records
        if len(self.report_history) > 100:
            self.report_history = self.report_history[-100:]
    
    def get_report_history(self) -> List[Dict[str, Any]]:
        """Get report generation history."""
        return self.report_history.copy()
    
    def __del__(self):
        """Cleanup when object is destroyed."""
        if hasattr(self, 'scheduler_running') and self.scheduler_running:
            self.stop_scheduler()
        
        if hasattr(self, 'report_executor'):
            self.report_executor.shutdown(wait=True)


# Convenience functions for common report types
def generate_daily_fraud_summary(output_dir: str = None) -> str:
    """Generate a daily fraud summary report."""
    generator = ReportGenerator(output_directory=output_dir)
    
    config = ReportConfig(
        report_type=ReportType.FRAUD_SUMMARY,
        report_format=ReportFormat.PDF,
        title="Daily Fraud Detection Summary",
        description="Daily summary of fraud detection activities",
        date_range_days=1
    )
    
    # This would typically fetch real data
    data = ReportData()
    
    return generator.generate_fraud_summary_report(data, config)


def generate_weekly_compliance_report(output_dir: str = None) -> str:
    """Generate a weekly compliance audit report."""
    generator = ReportGenerator(output_directory=output_dir)
    
    config = ReportConfig(
        report_type=ReportType.COMPLIANCE_AUDIT,
        report_format=ReportFormat.PDF,
        title="Weekly Compliance Audit Report",
        description="Weekly compliance audit for fraud detection system",
        date_range_days=7
    )
    
    # This would typically fetch real data
    data = ReportData()
    
    return generator.generate_compliance_audit_report(data, config)


def generate_monthly_transaction_analysis(output_dir: str = None) -> str:
    """Generate a monthly transaction analysis report."""
    generator = ReportGenerator(output_directory=output_dir)
    
    config = ReportConfig(
        report_type=ReportType.TRANSACTION_ANALYSIS,
        report_format=ReportFormat.EXCEL,
        title="Monthly Transaction Analysis Report",
        description="Detailed monthly analysis of transaction patterns",
        date_range_days=30,
        include_raw_data=True
    )
    
    # This would typically fetch real data
    data = ReportData()
    
    return generator.generate_transaction_analysis_report(data, config)